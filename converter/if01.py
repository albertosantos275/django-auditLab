import json 
import openpyxl
from . import if01_agregates 

def load_json(input_file):
    with open(input_file) as fp:
        data = json.load(fp)
        return data


person_types = None
operation_types = None
service_types = None
divisa_types = None
economic_activities = None
def load_references():
    global person_types , operation_types, service_types
    global divisa_types,economic_activities
    person_types = load_json('ref_person_types.json')
    operation_types = load_json('ref_operations.json')
    service_types = load_json('ref_services.json')
    divisa_types = load_json('ref_divisas.json')
    economic_activities = load_json('ref_activity_economic.json')

# def agregate_by_person_type(ref_sheet, min_col_number, max_col_number,row_number, codes): 
#     data={

#     }
#     risk_data={
#         "Alto":{"total":0},
#         "Bajo":{"total":0},
#         "Medio":{"total":0},
#     }
#     total = 0
#     risk_total =0
#     for k in codes.keys():
#         val = codes[k]
#         data[k]={
#             "total":0,
#             'user' : val['user'],
#             'risk' : val['risk']
#         }
#     for row in ref_sheet.iter_rows(min_row=row_number,
#                             max_col=max_col_number,
#                             min_col=min_col_number):
#         userType= row[0].value
#         transaction = row[9].value
#         monedad = row[10].value
        
#         if monedad == "DOP": 
#             if userType in data.keys():
#                 data[userType]["total"]+=transaction
#                 total+=transaction
#                 risk = codes[userType]['risk']
#                 if risk in risk_data.keys():
#                     risk_data[risk]['total']+=transaction
#                     risk_total+=transaction
#             else:
#                 print(f"user type not found {userType}")


#     for t in data.values():
#         t['perc']= t['total']/total*100

#     for t in risk_data.values():
#         t['perc']= t['total']/risk_total*100

        
#     temp={
#         "agregates" : data,
#         "total": total,
#         "risk_total":risk_total,
#         "risk": risk_data
#     } 
#     #TODO agregar total de las transacciones
#     # agregar cuenta de cada tipo   

#     return temp 


def compute_agregates(ref_sheet):
    
    by_person= if01_agregates.agregate_by_person_type(ref_sheet,2, 12,2, person_types)
    by_operation  = if01_agregates.agregate_by_operation(ref_sheet,1, 12,2, operation_types)
    by_service  = if01_agregates.agregate_by_services(ref_sheet,1, 12,2, service_types)
    by_activity  = if01_agregates.agregate_by_economic_activity(ref_sheet,1, 30,2, economic_activities)
    by_suspicious  = if01_agregates.agregate_by_suspicious_activity(ref_sheet,1, 30,2, economic_activities)

    agregates ={
        "by_person":by_person,
        "by_operations": by_operation,
        "by_service": by_service,
        "by_activity":by_activity,
        "by_suspicious":by_suspicious 
    }

    return agregates

    
def translate_row(ref_sheet, field, col_number, row_number, codes):
     for row in ref_sheet.iter_rows(min_row=row_number,
                            min_col=col_number,max_col=col_number):
        
        code  = row[0].value
        if code == None:
            continue
        code = str(code)
        
        if code in codes.keys():
            row[0].value= codes[code][field]
        else:
            print(code)

def translate(ref_sheet):
    translate_row(ref_sheet,'user',2, 2,person_types )
    translate_row(ref_sheet,'operation',7, 2,operation_types )
    translate_row(ref_sheet,'service',8, 2,service_types )
    translate_row(ref_sheet,'divisa',12, 2,divisa_types )
    translate_row(ref_sheet,'user',13, 2,person_types )
    translate_row(ref_sheet,'user',18, 2,person_types )
    translate_row(ref_sheet,'activity',27, 2,economic_activities )
    translate_row(ref_sheet,'activity',30, 2,economic_activities )
    translate_row(ref_sheet,'activity',33, 2,economic_activities )

        
def parse(input_file, output_file):
    load_references()
    workbook = openpyxl.load_workbook(input_file)
    ref_sheet = workbook.active
    agregates = compute_agregates(ref_sheet)
    translate(ref_sheet)

    output_file_agregates= output_file.split(".")[0]
    output_file_agregates+=".json"
    
    # with open(output_file_agregates,"w") as fp:
    #     json.dump(agregates, fp)
    workbook.save(output_file)
    return agregates

# r= parse('if1.xlsx', 'if1_m1.xlsx')
# print(r["by_suspicious"])

