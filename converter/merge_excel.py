
import glob
from openpyxl import load_workbook,Workbook

def createNewWorkbook(theOne,manyWb):
    dest = theOne.active
    first = True
    row_counter=1
    for wb in manyWb:
        source =   wb.active
        row_counter=copySheet(source,dest,row_counter, first)
        first= False

def copySheet(sourceSheet,newSheet,row_counter, first=False):
    current_row=0
    for row in sourceSheet.rows:
        current_row+=1
        if not first and current_row==1:
                continue
        for cell in row:
            newSheet.cell(row=row_counter, column=cell.col_idx,
                    value= cell.value)
        row_counter+=1
    return row_counter
       
            



def concat_files(paths, output_file):
    
    #path = "/scratch/excels"
    filesInput = paths
    list_of_files = [ load_workbook(f) for f in filesInput ]
    theOne = Workbook()
  
    createNewWorkbook(theOne,list_of_files)
    theOne.save(output_file)

# path="/scratch/excels"
# concat_files(path, path+ "/test.xlsx")