import json 
import openpyxl
from  . import fdo03b_agregates 

def load_json(input_file):
    with open(input_file) as fp:
        data = json.load(fp)
        return data


person_types = None
operation_types = None
service_types = None
divisa_types = None
economic_activities = None
localidades = None
payment_types = None
vinculacion_type = None
default_bank = None
sent_types = None
transaction_purpose = None
countries = None
cities = None
def load_references():
    global person_types , operation_types, service_types
    global divisa_types,economic_activities, localidades
    global payment_types, vinculacion_type, default_bank
    global sent_types, transaction_purposes, countries, cities

    
    person_types = load_json('ref_person_types.json')
    operation_types = load_json('ref_operations.json')
    service_types = load_json('ref_services.json')
    divisa_types = load_json('ref_divisas.json')
    # economic_activities = load_json('ref_activity_economic.json')
    localidades = load_json('ref_localidades.json')
    # payment_types = load_json('ref_pay_type.json')
    # vinculacion_type =  load_json('ref_vinculacion_type.json')
    # default_bank =   load_json('ref_default_bank.json')
    
    sent_types = load_json("ref_sent_type.json")
    transaction_purposes = load_json('ref_transaction_purpose.json')
    
    countries = load_json('ref_countries.json')
    cities = load_json ('ref_cities.json')

def compute_agregates(ref_sheet, filter_val):

   
    
    by_person= fdo03b_agregates.agregate_by_person_type(ref_sheet,12,8,0, 25,2, person_types, filter_val)
    # #ref_sheet, row_name, key_col, amount_col, 
    # #max_col_number,row_number, codes
    by_divisa = fdo03b_agregates.agregate_by_single_col(ref_sheet,'divisa',9,8,24,2, divisa_types, filter_val)
    by_cities = fdo03b_agregates.agregate_by_single_col(ref_sheet,'city',7,8,24,2, cities, filter_val)
    by_countries = fdo03b_agregates.agregate_by_single_col(ref_sheet,'country',6,8,24,2, countries, filter_val)
    by_purpose = fdo03b_agregates.agregate_by_single_col(ref_sheet,'name',17,8,24,2, transaction_purposes, filter_val)
    by_bank = fdo03b_agregates.agregate_by_single_col(ref_sheet,'name',3,8,24,2, {},filter_val,True)
    by_service  = fdo03b_agregates.agregate_by_services(ref_sheet,19,8,0, 25,2, service_types, filter_val)
    by_province  = fdo03b_agregates.agregate_by_provinces(ref_sheet,16,8,0, 25,2, localidades, filter_val)
    by_average  = fdo03b_agregates.agregate_by_average(ref_sheet,4,8,5,2,0,25,2, filter_val)

    # by_payment = ca01_agregates.agregate_by_single_col(ref_sheet,'name',15,3,24,2, payment_types)
    # by_vinculacion = ca01_agregates.agregate_by_single_col(ref_sheet,'name',21,3,24,2, vinculacion_type)
   
    # #by_operation  = ca01_agregates.agregate_by_operation(ref_sheet,0, 12,2, operation_types, default_bank)
    #                                                                 #user_col, amount_col,  min_col_number, max_col_number,row_number
    # by_activity  = ca01_agregates.agregate_by_economic_activity(ref_sheet,8,3,0, 25,2, economic_activities, default_bank)
    # by_suspicious  = ca01_agregates.agregate_by_suspicious_activity(ref_sheet,8,3,0,25,2, economic_activities, default_bank)

    agregates ={
        "by_person":by_person,
        "by_divisa":by_divisa,
        "by_cities":by_cities,
        "by_countries":by_countries,
        "by_purpose":by_purpose,
        "by_bank":by_bank,
        "by_service": by_service,
        "by_province": by_province,
        "by_average":by_average


        # "by_payment": by_payment,
        # "by_vinculacion" : by_vinculacion,
        # # "by_operations": by_operation,
        # "by_activity":by_activity,
        #  "by_suspicious":by_suspicious 
    }

    return agregates

    
def translate_row(ref_sheet, field, col_number, row_number, codes):
     for row in ref_sheet.iter_rows(min_row=row_number,
                            min_col=col_number,max_col=col_number):
        
        code  = row[0].value
        
        
        if code == None:
            continue
        code = str(code)
        
        if code in codes.keys():
            row[0].value= codes[code][field]
        else:
            print(code)

def translate(ref_sheet):
    translate_row(ref_sheet,'user',13, 2,person_types )
    translate_row(ref_sheet,'name',0, 2,sent_types )
    translate_row(ref_sheet,'name',18, 2,transaction_purposes )
    translate_row(ref_sheet,'country',7, 2,countries )
    translate_row(ref_sheet,'city',8, 2,cities )
    translate_row(ref_sheet,'nombre',17, 2,localidades )
    #translate_row(ref_sheet,'operation',7, 2,operation_types )
    translate_row(ref_sheet,'service',20, 2,service_types )
    #translate_row(ref_sheet,'divisa',24, 2,divisa_types )
    
   
    return

def create_sheet_by_filter(ref_sheet, output_file):


    return
        
def parse(input_file, output_file):

    


    load_references()
    workbook = openpyxl.load_workbook(input_file)
    ref_sheet = workbook.active
    #TODO , do preprocessing like converting to pesos
   
    agregates_r = compute_agregates(ref_sheet, 'R')
    agregates_s = compute_agregates(ref_sheet, 'E')
    agregates = {
        'recieved':agregates_r,
        'sent' :agregates_s
    }
    translate(ref_sheet)

    output_file_agregates= output_file.split(".")[0]
    output_file_agregates+=".json"
    
    # with open(output_file_agregates,"w") as fp:
    #     json.dump(agregates, fp)
    workbook.save(output_file)
    
    return agregates

# r= parse('fdo03b.xlsx', 'fdo3b_m.xlsx')
# print(r['sent']['by_average'])
# with open("test.json","w") as fp:
#     json.dump(r['by_bank'], fp)

