

import json

import openpyxl

loc = ("./tabla_refs_completa.xlsx")
workbook = openpyxl.load_workbook(loc)

ref_sheet = workbook["Tablas de Apoyo"]

# Give the location of the file
def dump_json(data, filename):
    with open(filename, "w") as fp:
        json.dump(data, fp)


def ref_person():
    data= {}
    #tipo de persona
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=2,max_col=4):
        codigo = row[0].value
        user = row[1].value
        risk  = row[2].value
        if codigo == None or user == None or risk == None:
            continue
        data[codigo] = {
            'codigo': codigo,
            'user': user,
            'risk': risk
        }                    
    dump_json(data, "ref_person_types.json")

def ref_code_country():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=7,max_col=8):
        codigo = row[0].value
        country = row[1].value
        if codigo == None or country == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'country': country,
        }                    
    dump_json(data, "ref_countries.json")


# ref_code_country()

def ref_code_divisas():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=14,max_col=15):
        codigo = row[0].value
        divisa = row[1].value
        if codigo == None or divisa == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'divisa': divisa,
        }                    
    dump_json(data, "ref_divisas.json")


def ref_code_operation():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=18,max_col=19):
        codigo = row[0].value
        operation = row[1].value
       
        if codigo == None or operation == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'operation': operation,
        }                    
    dump_json(data, "ref_operations.json")

def ref_code_services():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=22,max_col=24):
        codigo = row[0].value
        service = row[1].value
        fase = row[2].value
        if codigo == None and service !=None:
            tipo = service
            continue
        
       
        if codigo == None or service == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'service': service,
            'tipo' : tipo,
            "phase": fase
        }                    
    dump_json(data, "ref_services.json")
# ref_code_operation()


#ref_code_services()

def ref_code_actividad_economica():
    data={}
    count=0
    for row in ref_sheet.iter_rows(min_row=2,
                            min_col=4,max_col=12):
        if count>=2851:
            break
        else:
            count+=1
        codigo = row[0].value
        macro = row[1].value
        activity = row[2].value
        p1= row[3].value
        p2= row[4].value
        p3= row[5].value
        r1= row[6].value
        r2=row[7].value
        r3=row[8].value
        if codigo == None or activity == None:
                continue
        codigo = str(codigo)
        activity = activity.replace(codigo, '').strip(' ')
        macro = str(macro)
        macro = macro.strip(",")
        
        data[codigo] = {
            'codigo': codigo,
            'macro': macro,
            'activity' : activity,
            "Delito Precedente 1":p1, 
            "Delito Precedente 2":p2,
            "Delito Precedente 3":p3,
            'Riesgo 1': r1,
            "Riesgo 2":r2,
            "Riesgo 3":r3,
        } 
        # print(data)
        # break                   
    dump_json(data, "ref_activity_economic.json")

# loc = ("./ref_table_2.xlsx")
# workbook = openpyxl.load_workbook(loc)

# ref_sheet = workbook["CIIU VF"]
# ref_code_actividad_economica()


def ref_code_localidades():
    data={}
    count=0
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=0,max_col=13):
        
        region = str(row[0].value)
        provincia = row[1].value
        municipio = row[2].value
        district = row[3].value
        seccion = row[4].value
        barrio = row[5].value
        sub_barrio = row[6].value
        codigo = row[7].value
        name = row[8].value
        economic = row[9].value
        p1 = row[10].value
        p2 = row[11].value
        p3 = row[12].value
        codigo = str(codigo)
        codigo = codigo.strip(' ')
        diff = 6- len(codigo)
        if  diff> 0:
            s='0'*diff
            codigo = s+codigo
        
        # activity = activity.replace(codigo, '').strip(' ')
        # macro = str(macro)
        # macro = macro.strip(",")
        
        data[codigo] = {
            'codigo': codigo,
            "Delito Precedente 1":p1, 
            "Delito Precedente 2":p2,
            "Delito Precedente 3":p3,
            "nombre" : name,
            "region" : region,
            "provincia" : provincia,
            "district": district,
            "seccion": seccion,
            "barrio" : barrio,
            "economic": economic,
            "subbarrio" : sub_barrio,
            "municipio" : municipio
        } 
                         
    dump_json(data, "ref_localidades.json")

# loc = ("./localidades.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tabla 16 "]
# ref_code_localidades()


def ref_payment_types():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=19,max_col=25):
        codigo = row[0].value
        divisa = row[1].value
        # print(codigo)
        # print(divisa)
        # break
        if codigo == None or divisa == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'name': divisa,
        }                    
    dump_json(data, "ref_pay_type.json")


def ref_vinculacion_types():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=25,max_col=30):
        codigo = str(row[0].value).strip(' ')
        divisa = str(row[1].value).strip(' ')
        # print(codigo)
        # print(divisa)
        # break
        if codigo == None or divisa == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'name': divisa,
        }                    
    dump_json(data, "ref_vinculacion_type.json")

# loc = ("./ca01_refs.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tablas de Apoyo"]

# ref_vinculacion_types()
# ref_payment_types()

def ref_default_banco():
    data={}
    for row in ref_sheet.iter_rows(min_row=2,
                            min_col=0,max_col=30):
        id_client = row[0].value                    
        codigo = row[1].value
        name = row[2].value
        last_name = row[3].value
        ciiu = str(row[4].value)
        ciuu = ciiu.strip(' ')
        
        # print(codigo)
        # print(divisa)
        # break
        if codigo == None:
                continue
        data[codigo] = {
            'id': id_client,
            'codigo': codigo,
            'name': name,
            'lastname' : last_name,
            'ciiu': ciiu
        }                    
    dump_json(data, "ref_default_bank.json")


# loc = ("./banco_d.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Sheet1"]
# ref_default_banco()



def ref_transaction_purpose():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=0,max_col=30):
                     
        codigo = row[18].value
        name = row[19].value
        # print(f'{codigo} {name}')
        # break
        
        # print(codigo)
        # print(divisa)
        # break
        if codigo == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'name': name,
            
        }                    
    dump_json(data, "ref_transaction_purpose.json")


# loc = ("./FD03B_ref.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tablas de Apoyo"]
# ref_transaction_purpose()



def ref_cities():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=0,max_col=30):
                     
        codigo = str(row[22].value)
        name = row[23].value
        diff = 7-len(codigo)
        if diff>0:
            codigo = '0'*diff+codigo

        # print(f'{codigo} {name}')
        # break
        
        # print(codigo)
        # print(divisa)
        # break
        if codigo == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'city': name,
            
        }                    
    dump_json(data, "ref_cities.json")


# loc = ("./FD03B_ref.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tablas de Apoyo"]
# ref_cities()


def ref_origin():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=0,max_col=40):
                     
        codigo = str(row[14].value)
        name = row[15].value
        # diff = 7-len(codigo)
        # if diff>0:
        #     codigo = '0'*diff+codigo

        # print(f'{codigo} {name}')
        # break
        
        if codigo == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'name': name,
            
        }                    
    dump_json(data, "ref_origin_resources.json")


# loc = ("./de11_ref.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tablas de Apoyo"]
# ref_origin()


def ref_interes_way():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=0,max_col=40):
                     
        codigo = str(row[10].value)
        name = row[11].value
        # diff = 7-len(codigo)
        # if diff>0:
        #     codigo = '0'*diff+codigo

        print(f'{codigo} {name}')
        #break
        
        if codigo == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'name': name,
            
        }                    
    dump_json(data, "ref_interes_way.json")


# loc = ("./de11_ref.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tablas de Apoyo"]
# ref_interes_way()


def client_type():
    data={}
    for row in ref_sheet.iter_rows(min_row=4,
                            min_col=0,max_col=40):
                     
        codigo = str(row[24].value)
        name = row[25].value
        definition = row[26].value
        r1 = row[27].value
        # diff = 7-len(codigo)
        # if diff>0:
        #     codigo = '0'*diff+codigo

        # print(f'{codigo} {name} {definition} {r1}')
        # break
        
        if codigo == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'type': name,
            'definition':definition,
            'risk': r1
            
        }                    
    dump_json(data, "ref_client_types.json")


# loc = ("./de11_ref.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tablas de Apoyo"]
# client_type()


def ref_flex_normativa():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=0,max_col=40):
                     
        codigo = str(row[30].value)
        name = row[31].value
        # diff = 7-len(codigo)
        # if diff>0:
        #     codigo = '0'*diff+codigo

        # print(f'{codigo} {name}')
        # break
        
        if codigo == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'name': name,
            
        }                    
    dump_json(data, "ref_flex_normativa.json")


# loc = ("./de11_ref.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tablas de Apoyo"]
# ref_flex_normativa()


def ref_credit_operation():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=0,max_col=40):
                     
        codigo = str(row[34].value)
        name = row[35].value
        # diff = 7-len(codigo)
        # if diff>0:
        #     codigo = '0'*diff+codigo

        # print(f'{codigo} {name}')
        # break
        
        if codigo == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'name': name,
            
        }                    
    dump_json(data, "ref_credit_operation.json")


# loc = ("./de11_ref.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tablas de Apoyo"]
# ref_credit_operation()


def ref_classification_reason():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=0,max_col=40):
                     
        codigo = str(row[38].value)
        name = row[39].value
        # diff = 7-len(codigo)
        # if diff>0:
        #     codigo = '0'*diff+codigo

        # print(f'{codigo} {name}')
        # break
        
        if codigo == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'name': name,
            
        }                    
    dump_json(data, "ref_classification_reason.json")


# loc = ("./de11_ref.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tablas de Apoyo"]
# ref_classification_reason()



def ref_buy_source():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=0,max_col=40):
                     
        codigo = str(row[14].value)
        name = row[15].value
        
        if codigo == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'name': name,
            
        }                    
    dump_json(data, "ref_buy_source.json")


# loc = ("./fd01_ref.xlsx")
# workbook = openpyxl.load_workbook(loc)
# ref_sheet = workbook["Tablas de Apoyo"]
# ref_buy_source()



def ref_pay_way():
    data={}
    for row in ref_sheet.iter_rows(min_row=3,
                            min_col=0,max_col=40):
                     
        codigo = str(row[18].value)
        name = row[19].value
        # print(f"{codigo} {name}")
        # break
        if codigo == None:
                continue
        data[codigo] = {
            'codigo': codigo,
            'name': name,
            
        }                    
    dump_json(data, "ref_pay_way.json")


loc = ("./fd01_ref.xlsx")
workbook = openpyxl.load_workbook(loc)
ref_sheet = workbook["Tablas de Apoyo"]
ref_pay_way()